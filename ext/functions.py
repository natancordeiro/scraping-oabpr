"""
Módulo onde é guardado as funções externas para utilização geral do projeto.
"""

import os
import pydub
import urllib
import random
import logging
import colorlog
import speech_recognition
from datetime import datetime
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def setup_logging(to_file=False):
    """Setup logging"""

    # Criar um logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Formato do log
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    date_format = '%d-%m-%Y %H:%M:%S'
    formatter = logging.Formatter(log_format, datefmt=date_format)

    # Configurar o handler para o console com cores
    color_formatter = colorlog.ColoredFormatter(
        '%(log_color)s' + log_format,
        datefmt=date_format,
        reset=True,
        log_colors={
            'DEBUG': 'white',
            'INFO': 'green',
            'WARNING': 'yellow',
            'ERROR': 'red',
            'CRITICAL': 'bold_red',
        }
    )
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(color_formatter)
    logger.addHandler(console_handler)

    if to_file:
        # Criar pasta logs se não existir
        if not os.path.exists('logs'):
            os.makedirs('logs')
        
        # Nome do arquivo de log com a data e hora atuais
        log_filename = datetime.now().strftime("logs/log_%d-%m-%Y_%H-%M-%S.log")
        
        # Configurar o handler para o arquivo de log
        file_handler = RotatingFileHandler(log_filename, maxBytes=10**6, backupCount=5)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

    return logger

def substituir_ultima_letra(palavra: str) -> str:
    """
    Substitui a última letra de uma palavra.

    Args:
        palavra (str): Palavra a ser modificada.
    """
    
    if palavra.endswith("a"):
        return palavra[:-1] + "o"
    return palavra

def get_temps_files() -> tuple:
    """Retorna os caminhos para os arquivos temporários"""

    path_to_mp3 = os.path.normpath(os.path.join((os.getenv("TEMP") if os.name=="nt" else "/tmp/")+ str(random.randrange(1,1000))+".mp3"))
    path_to_wav = os.path.normpath(os.path.join((os.getenv("TEMP") if os.name=="nt" else "/tmp/")+ str(random.randrange(1,1000))+".wav"))
    return path_to_mp3, path_to_wav
        
def convert_audio_to_string(audio_source, path_to_mp3, path_to_wav):
    """Converte o arquivo de áudio para texto"""

    urllib.request.urlretrieve(audio_source, path_to_mp3)

    os.environ["PATH"] += os.pathsep + 'C:\\ffmpeg'
    sound = pydub.AudioSegment.from_mp3(path_to_mp3)
    sound.export(path_to_wav, format="wav")
    sample_audio = speech_recognition.AudioFile(path_to_wav)
    r = speech_recognition.Recognizer()
    with sample_audio as source:
        audio = r.record(source)

    key = r.recognize_google(audio)
    os.remove(path_to_mp3)
    os.remove(path_to_wav)
    return key

def get_excel_filename():
    """Retorna o caminho para o arquivo de saída"""

    filename = os.path.join(os.getcwd(), "advogados_OABPR.xlsx")
    return filename

def create_excel_file():
    """
    Cria um arquivo excel vazio.
    
    :param filename: Nome do arquivo a ser criado.
    """
    
    filename = get_excel_filename()
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        headers = ["Número de Inscrição", "Advogado", "Impedimentos", "Situação", "Subseção", "Data da Inscrição", "Endereço Comercial", "Telefone Comercial"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        wb.save(filename)
        return True
    else:
        return False

def insert_values(values: list):
    """
    Insere os dados processados na planilha de saída.
    
    :param values: Dados a serem inseridos na planilha.
    """

    filename = get_excel_filename()
    wb = load_workbook(filename)
    ws = wb.active
    row = ws.max_row + 1
    ws.append(values)
    wb.save(filename)
